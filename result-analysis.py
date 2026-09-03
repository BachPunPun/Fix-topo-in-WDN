# -*- coding: utf-8 -*-
import sys, os
from pathlib import Path

# Thêm đường dẫn thư mục scripts, function và thư mục gốc
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "scripts"))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "function"))
sys.path.insert(0, os.path.dirname(__file__))

import setup_console
import importlib
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import geopandas as gpd

from load_layers import load_layers
import config

# Import pipeline và helper từ module atopo
mod_atopo = importlib.import_module("atopo")

# ── Cấu hình thư mục và danh sách tham số ─────────────────────
OUTPUT_DIR = config.OUTPUT_DIR / "result-analysis"
RESULT_EXCEL = OUTPUT_DIR / "result.xlsx"

TOLERANCE_VALUES = [0.001, 0.005, 0.01, 0.05, 0.1, 0.15, 0.2, 0.25, 0.3, 0.35, 0.4, 0.45, 0.5]


# ── Hàm chạy Atopo cho một giá trị tolerance ──────────────────
def run_atopo_for_tolerance(gdf_orig: gpd.GeoDataFrame, tol: float):
    """
    Chạy toàn bộ pipeline Atopo với giá trị tolerance xác định trên bản sao dữ liệu.
    Trả về (topo_report, network_report, components, gdf).
    """
    gdf = gdf_orig.copy(deep=True)
    topo_report = []
    network_report = []

    # Bước khởi tạo thống kê mạng lưới
    network_report.append(("Bắt đầu", mod_atopo._calc_network_stats(gdf)))

    # Chạy tuần tự các bước trong PIPELINE
    for step_name, _file_suffix, step_fn in mod_atopo.PIPELINE:
        gdf, detect_cnt, fixed_cnt, skip_cnt, step_errors = step_fn(gdf, tolerance=tol)
        topo_report.append((step_name, detect_cnt, fixed_cnt, skip_cnt))
        network_report.append((step_name, mod_atopo._calc_network_stats(gdf)))

    # Lấy thông tin chi tiết các nhóm thành phần liên thông
    components = mod_atopo._get_cc_detail(gdf)

    return topo_report, network_report, components, gdf


# ── Hàm ghi định dạng dữ liệu vào sheet openpyxl ──────────────
def write_sheet_content(ws, tol: float, topo_report: list, network_report: list, components: list):
    """
    Ghi cả 3 bảng báo cáo vào worksheet:
      1. Báo cáo lỗi topology
      2. Sự thay đổi thành phần liên thông lớn
      3. Nhóm thành phần liên thông (sắp xếp giảm dần theo số edges)
    Kèm định dạng kẻ bảng, màu nền header, căn lề và chỉnh độ rộng cột.
    """
    # Font & Style
    title_font = Font(name="Calibri", size=12, bold=True, color="1F4E79")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)

    header_fill_1 = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_fill_2 = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_fill_3 = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    total_border = Border(
        left=Side(style='thin', color='A6A6A6'),
        right=Side(style='thin', color='A6A6A6'),
        top=Side(style='thin', color='A6A6A6'),
        bottom=Side(style='double', color='000000')
    )

    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_right = Alignment(horizontal='right', vertical='center')

    current_row = 1

    # ══════════════════════════════════════════════════════════
    # 1. BÁO CÁO LỖI TOPOLOGY
    # ══════════════════════════════════════════════════════════
    ws.cell(row=current_row, column=1, value=f"1. Báo cáo lỗi topology (tolerance = {tol} m)").font = title_font
    current_row += 2

    # Headers bảng 1
    headers_1 = ["STT", "Type", "Detect", "Correct", "Skip"]
    for col_idx, h in enumerate(headers_1, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill_1
        cell.alignment = align_center
        cell.border = thin_border
    current_row += 1

    total_d, total_c, total_s = 0, 0, 0
    for i, (name, d, c, s) in enumerate(topo_report, start=1):
        total_d += d
        total_c += c
        total_s += s

        row_data = [i, name, d, c, s]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = align_center
            elif col_idx == 2:
                cell.alignment = align_left
            else:
                cell.alignment = align_right
        current_row += 1

    # Dòng Total bảng 1
    total_row_1 = ["", "Total", total_d, total_c, total_s]
    for col_idx, val in enumerate(total_row_1, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=val)
        cell.font = bold_font
        cell.fill = total_fill
        cell.border = total_border
        if col_idx == 2:
            cell.alignment = align_left
        else:
            cell.alignment = align_right if col_idx > 2 else align_center
    current_row += 3

    # ══════════════════════════════════════════════════════════
    # 2. SỰ THAY ĐỔI THÀNH PHẦN LIÊN THÔNG LỚN
    # ══════════════════════════════════════════════════════════
    ws.cell(row=current_row, column=1, value=f"2. Sự thay đổi thành phần liên thông lớn (tolerance = {tol} m)").font = title_font
    current_row += 2

    # Headers bảng 2
    headers_2 = [
        "STT",
        "Bước",
        "Tổng cạnh",
        "Tổng cạnh trong\nthành phần lớn",
        "Thành phần\nliên thông",
        "CC Δ%",
        "Tổng chiều dài",
        "Tổng chiều dài trong\nthành phần lớn",
        "Tỷ lệ chiều dài của\nthành phần lớn (%)"
    ]
    for col_idx, h in enumerate(headers_2, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill_2
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[current_row].height = 30
    current_row += 1

    cc_start = network_report[0][1]['CC']
    for i, (name, s) in enumerate(network_report, start=1):
        if i == 1:
            cc_delta = "—"
        else:
            pct = (cc_start - s['CC']) / cc_start * 100 if cc_start else 0
            cc_delta = f"{pct:+.1f}%"

        row_data = [
            i,
            name,
            s['Edges'],
            s['LargestEdges'],
            s['CC'],
            cc_delta,
            round(s['TotalLen'], 2),
            round(s['LargestLen'], 2),
            round(s['LenRatio'], 2)
        ]

        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = align_center
            elif col_idx == 2:
                cell.alignment = align_left
            elif col_idx == 6:
                cell.alignment = align_center
            else:
                cell.alignment = align_right
                if col_idx in (7, 8):
                    cell.number_format = '#,##0.00'
                elif col_idx == 9:
                    cell.number_format = '0.00'
        current_row += 1

    # Dòng Final bảng 2
    final = network_report[-1][1]
    pct_final = (cc_start - final['CC']) / cc_start * 100 if cc_start else 0
    cc_delta_final = f"{pct_final:+.1f}%"

    final_row_2 = [
        "",
        "Final",
        final['Edges'],
        final['LargestEdges'],
        final['CC'],
        cc_delta_final,
        round(final['TotalLen'], 2),
        round(final['LargestLen'], 2),
        round(final['LenRatio'], 2)
    ]
    for col_idx, val in enumerate(final_row_2, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=val)
        cell.font = bold_font
        cell.fill = total_fill
        cell.border = total_border
        if col_idx == 2:
            cell.alignment = align_left
        elif col_idx == 6:
            cell.alignment = align_center
        else:
            cell.alignment = align_right if col_idx > 2 else align_center
            if col_idx in (7, 8):
                cell.number_format = '#,##0.00'
            elif col_idx == 9:
                cell.number_format = '0.00'
    current_row += 3

    # ══════════════════════════════════════════════════════════
    # 3. NHÓM THÀNH PHẦN LIÊN THÔNG (SẮP XẾP GIẢM DẦN THEO SỐ EDGES)
    # ══════════════════════════════════════════════════════════
    ws.cell(row=current_row, column=1, value=f"3. Nhóm thành phần liên thông (sắp xếp giảm dần theo số edges) (tolerance = {tol} m)").font = title_font
    current_row += 2

    # Headers bảng 3
    headers_3 = ["STT", "Nhóm", "Số cạnh (Edges)", "Danh sách ID"]
    for col_idx, h in enumerate(headers_3, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill_3
        cell.alignment = align_center
        cell.border = thin_border
    current_row += 1

    if components:
        # Nhóm 1 (Thành phần lớn nhất)
        stt = 1
        edges_main = components[0][0]
        row_data = [stt, "Nhóm 1", edges_main, "—"]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = align_center
            elif col_idx == 2:
                cell.alignment = align_left
            elif col_idx == 3:
                cell.alignment = align_right
                cell.number_format = '#,##0'
            else:
                cell.alignment = align_center
        current_row += 1
        stt += 1

        # Các nhóm có >= 2 edges
        group_num = 2
        single_edge_ids = []
        for edges_count, ids in components[1:]:
            if edges_count >= 2:
                id_str = ", ".join(ids)
                if len(id_str) > 32000:
                    id_str = id_str[:32000] + "... (truncated)"
                row_data = [stt, f"Nhóm {group_num}", edges_count, id_str]
                for col_idx, val in enumerate(row_data, start=1):
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    cell.font = regular_font
                    cell.border = thin_border
                    if col_idx == 1:
                        cell.alignment = align_center
                    elif col_idx == 2:
                        cell.alignment = align_left
                    elif col_idx == 3:
                        cell.alignment = align_right
                        cell.number_format = '#,##0'
                    else:
                        cell.alignment = align_left_wrap
                current_row += 1
                group_num += 1
                stt += 1
            else:
                single_edge_ids.extend(ids)

        # Gom nhóm các thành phần chỉ có 1 edge
        if single_edge_ids:
            num_single = sum(1 for ec, _ in components[1:] if ec == 1)
            id_str = ", ".join(single_edge_ids)
            if len(id_str) > 32000:
                id_str = id_str[:32000] + "... (truncated)"
            row_data = [stt, f"Nhóm 1 edges ({num_single})", num_single, id_str]
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = regular_font
                cell.border = thin_border
                if col_idx == 1:
                    cell.alignment = align_center
                elif col_idx == 2:
                    cell.alignment = align_left
                elif col_idx == 3:
                    cell.alignment = align_right
                    cell.number_format = '#,##0'
                else:
                    cell.alignment = align_left_wrap
            current_row += 1
            stt += 1

    # Dòng Total bảng 3
    total_edges_cc = sum(c[0] for c in components) if components else 0
    total_row_3 = ["", "Total", total_edges_cc, f"{len(components)} thành phần liên thông"]
    for col_idx, val in enumerate(total_row_3, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=val)
        cell.font = bold_font
        cell.fill = total_fill
        cell.border = total_border
        if col_idx == 2:
            cell.alignment = align_left
        elif col_idx == 3:
            cell.alignment = align_right
            cell.number_format = '#,##0'
        elif col_idx == 4:
            cell.alignment = align_left
        else:
            cell.alignment = align_center

    # ── Tự động điều chỉnh độ rộng cột ────────────────────────
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            val_str = str(cell.value or "")
            # Bỏ qua dòng tiêu đề dài khi tính width cột A
            if col[0].column == 1 and val_str.startswith(("1. ", "2. ", "3. ")):
                continue
            if "\n" in val_str:
                lines = val_str.split("\n")
                line_max = max(len(l) for l in lines)
                max_len = max(max_len, line_max)
            else:
                max_len = max(max_len, len(val_str))

        col_width = max(max_len + 4, 12)
        # Giới hạn bề rộng cho cột D (Danh sách ID) để không bị quá rộng
        if col_letter == 'D':
            col_width = min(col_width, 60)
        ws.column_dimensions[col_letter].width = col_width


# ── Hàm chạy chính ────────────────────────────────────────────
def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    print("=" * 75)
    print("🚀 BẮT ĐẦU CHẠY PHÂN TÍCH KẾT QUẢ ATOPO THEO CÁC MỨC TOLERANCE")
    print(f"📌 Danh sách Tolerance: {TOLERANCE_VALUES}")
    print(f"📁 Thư mục lưu kết quả: {OUTPUT_DIR}")
    print("=" * 75)

    print("\nĐang nạp dữ liệu lớp ống...")
    gdf_orig, layer_meta = load_layers()
    print(f"→ Đã nạp {len(gdf_orig)} ống từ {len(layer_meta)} lớp: {', '.join(layer_meta)}")

    # Tạo workbook mới
    wb = openpyxl.Workbook()
    # Xóa sheet mặc định ban đầu
    default_sheet = wb.active

    for idx, tol in enumerate(TOLERANCE_VALUES, start=1):
        sheet_name = str(tol)
        print(f"\n[{idx}/{len(TOLERANCE_VALUES)}] Đang chạy Atopo với Tolerance = {tol} m ...")

        topo_report, network_report, components, gdf_out = run_atopo_for_tolerance(gdf_orig, tol)

        total_d = sum(r[1] for r in topo_report)
        total_c = sum(r[2] for r in topo_report)
        cc_final = network_report[-1][1]['CC']
        len_ratio = network_report[-1][1]['LenRatio']

        print(f"  • Lỗi phát hiện: {total_d:>4} | Đã sửa: {total_c:>4}")
        print(f"  • Thành phần liên thông (CC): {cc_final:>3} | Tỷ lệ chiều dài thành phần lớn: {len_ratio:.2f}%")
        print(f"  • Số nhóm liên thông: {len(components)}")

        ws = wb.create_sheet(title=sheet_name)
        write_sheet_content(ws, tol, topo_report, network_report, components)

    # Xóa sheet mặc định trống
    if default_sheet and default_sheet in wb.worksheets and len(wb.worksheets) > 1:
        wb.remove(default_sheet)

    # Lưu workbook
    wb.save(RESULT_EXCEL)

    print("\n" + "═" * 75)
    print("🎉 HOÀN THÀNH TẤT CẢ PHÂN TÍCH KẾT QUẢ ATOPO!")
    print(f"✅ File Excel kết quả đã lưu tại: {RESULT_EXCEL}")
    print("═" * 75)


if __name__ == "__main__":
    main()
